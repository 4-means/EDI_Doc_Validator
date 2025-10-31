import argparse
import csv
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from typing import List, Tuple, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

EDI_FORMAT_X12 = "X12"
EDI_FORMAT_EDIFACT = "EDIFACT"


def detect_edi_format(raw: str) -> str:
    stripped = raw.lstrip()
    if stripped.startswith("UNA") or stripped.startswith("UNB"):
        return EDI_FORMAT_EDIFACT
    return EDI_FORMAT_X12


def parse_una(raw: str) -> Dict[str, Optional[str]]:
    idx = raw.find("UNA")
    if idx == -1:
        return {
            "component_separator": ":",
            "element_separator": "+",
            "decimal_mark": ".",
            "release_character": "?",
            "repetition_separator": None,
            "segment_terminator": "'"
        }
    data = raw[idx + 3: idx + 9]
    if len(data) < 6:
        return {
            "component_separator": ":",
            "element_separator": "+",
            "decimal_mark": ".",
            "release_character": "?",
            "repetition_separator": None,
            "segment_terminator": "'"
        }
    return {
        "component_separator": data[0],
        "element_separator": data[1],
        "decimal_mark": data[2],
        "release_character": data[3],
        "repetition_separator": data[4] if data[4] != " " else None,
        "segment_terminator": data[5]
    }


def split_segments(raw: str, terminator: str, release: Optional[str] = None) -> List[str]:
    segments: List[str] = []
    current: List[str] = []
    i = 0
    length = len(raw)
    while i < length:
        ch = raw[i]
        if release and ch == release:
            i += 1
            if i < length:
                current.append(raw[i])
            i += 1
            continue
        if ch == terminator:
            seg = "".join(current).strip()
            if seg:
                segments.append(seg)
            current = []
        else:
            if ch not in ("\r", "\n"):
                current.append(ch)
        i += 1
    tail = "".join(current).strip()
    if tail:
        segments.append(tail)
    return segments


def split_elements(segment: str, element_sep: str, release: Optional[str] = None) -> List[str]:
    if not element_sep:
        return [segment]
    elems: List[str] = []
    current: List[str] = []
    i = 0
    length = len(segment)
    while i < length:
        ch = segment[i]
        if release and ch == release:
            i += 1
            if i < length:
                current.append(segment[i])
            i += 1
            continue
        if ch == element_sep:
            elems.append("".join(current))
            current = []
        else:
            current.append(ch)
        i += 1
    elems.append("".join(current))
    return elems


# -------------------------------------------------------
#  File I/O and parsing
# -------------------------------------------------------

def read_edi_file(path: str,
                  seg_term: Optional[str] = None,
                  elem_sep_override: Optional[str] = None) -> Tuple[List[str], Dict[str, Any]]:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        raw = f.read()
    edi_format = detect_edi_format(raw)

    if edi_format == EDI_FORMAT_EDIFACT:
        una = parse_una(raw)
        release_char = una.get("release_character")
        seg_sep = seg_term or una.get("segment_terminator") or "'"
        element_sep = una.get("element_separator") or "+"
        component_sep = una.get("component_separator") or ":"
        repetition_sep = una.get("repetition_separator")
        segments = split_segments(raw.replace("\r", "").replace("\n", ""), seg_sep, release_char)
        meta = {
            "format": edi_format,
            "segment_terminator": seg_sep,
            "element_separator": element_sep,
            "component_separator": component_sep,
            "release_character": release_char,
            "repetition_separator": repetition_sep
        }
        return segments, meta

    guessed = "~" if "~" in raw else None
    seg_sep = seg_term or guessed or "~"
    element_sep = elem_sep_override or "*"
    segments = split_segments(raw.replace("\r", "").replace("\n", ""), seg_sep, None)
    meta = {
        "format": edi_format,
        "segment_terminator": seg_sep,
        "element_separator": element_sep,
        "component_separator": None,
        "release_character": None,
        "repetition_separator": None
    }
    return segments, meta


def read_x12(path: str, seg_term: str = None) -> Tuple[List[str], str]:
    segments, meta = read_edi_file(path, seg_term=seg_term, elem_sep_override="*")
    return segments, meta["segment_terminator"]


def parse_segments(segments: List[str],
                   element_sep: str = "*",
                   release_char: Optional[str] = None) -> List[Dict[str, Any]]:
    parsed = []
    for idx, seg in enumerate(segments):
        elems = split_elements(seg, element_sep, release_char)
        tag = elems[0] if elems else ""
        parsed.append({"index": idx, "tag": tag, "elements": elems})
    return parsed


# -------------------------------------------------------
#  Detect ST01 (tx type) from the pattern file
# -------------------------------------------------------
def detect_tx_type(parsed: List[Dict[str, Any]], format_hint: Optional[str] = None) -> Optional[str]:
    if format_hint in (None, EDI_FORMAT_EDIFACT):
        for rec in parsed:
            if rec["tag"] == "UNH":
                if len(rec["elements"]) > 2 and rec["elements"][2]:
                    return rec["elements"][2].split(":")[0]
    for rec in parsed:
        if rec["tag"] == "ST":
            return rec["elements"][1] if len(rec["elements"]) > 1 else None
    return None


def normalize_tx_type(tx_type: Optional[str], format_hint: str) -> Optional[str]:
    if not tx_type:
        return None
    canonical = tx_type.upper()
    if format_hint == EDI_FORMAT_EDIFACT:
        mapping = {
            "ASN": "DESADV",
            "DESADV": "DESADV",
            "QUALITY": "QALITY",
            "QALITY": "QALITY",
            "QUALITYREPORT": "QALITY"
        }
        return mapping.get(canonical, canonical)
    mapping = {
        "ASN": "856",
        "DESADV": "856",
        "QUALITY": "863",
        "QALITY": "863"
    }
    return mapping.get(canonical, canonical)


# -------------------------------------------------------
#  Ignore rules (CSV)
# -------------------------------------------------------
class IgnoreRule:
    __slots__ = ("segment_tag", "elem", "qual_pos", "qual_eq", "comp_idx")

    def __init__(self, segment_tag: str, elem: str,
                 qual_pos: Optional[int], qual_eq: Optional[str], comp_idx: Optional[int]):
        self.segment_tag = (segment_tag or "").strip()
        self.elem = (elem or "").strip()  # "*" or numeric string
        self.qual_pos = qual_pos
        self.qual_eq = (qual_eq or None)
        self.comp_idx = comp_idx  # not used for simple X12 elements


def load_ignore_rules(path: Optional[str]) -> List[IgnoreRule]:
    rules: List[IgnoreRule] = []
    if not path:
        return rules
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            seg = (row.get("segment_tag") or "").strip()
            elem = (row.get("element_index_raw") or "").strip()
            qp_raw = (row.get("qualifier_pos") or "").strip()
            qe = (row.get("qualifier_equals") or "").strip()
            ci_raw = (row.get("component_index") or "").strip()
            qual_pos = int(qp_raw) if qp_raw.isdigit() else None
            comp_idx = int(ci_raw) if ci_raw.isdigit() else None
            rules.append(IgnoreRule(seg, elem, qual_pos, qe if qe else None, comp_idx))
    return rules


def is_ignored(rules: List[IgnoreRule], seg_tag: str, elem_idx_raw: int, elems_for_qual: List[str]) -> bool:
    if not rules:
        return False
    for r in rules:
        if r.segment_tag != seg_tag:
            continue
        if not (r.elem == "*" or (r.elem.isdigit() and int(r.elem) == elem_idx_raw)):
            continue
        if r.qual_pos is not None:
            if r.qual_pos >= len(elems_for_qual):
                continue
            if r.qual_eq is not None and elems_for_qual[r.qual_pos] != r.qual_eq:
                continue
        return True
    return False


# -------------------------------------------------------
#  Element comparison (adds issue + context columns)
# -------------------------------------------------------
def compare_elements(p_seg: Dict[str, Any], t_seg: Dict[str, Any],
                     ignore_rules: List[IgnoreRule],
                     line_key: str = "", cid_key: str = "") -> Tuple[List[Dict[str, Any]], bool]:
    rows = []
    any_diff = False

    p_elems = p_seg["elements"]
    t_elems = t_seg["elements"]
    maxlen = max(len(p_elems), len(t_elems))

    for i in range(maxlen):
        p_val = p_elems[i] if i < len(p_elems) else None
        t_val = t_elems[i] if i < len(t_elems) else None

        status = "OK"
        if p_val is None and t_val is not None:
            status = "EXTRA_ELEMENT"
        elif p_val is not None and t_val is None:
            status = "MISSING_ELEMENT"
        elif p_val != t_val:
            status = "DIFF" if i > 0 else "TAG_DIFF"

        ignored = is_ignored(ignore_rules, t_seg["tag"] or p_seg["tag"], i, t_elems)
        status_effective = "IGNORED" if ignored else status

        if not ignored and status in ("EXTRA_ELEMENT", "MISSING_ELEMENT", "DIFF", "TAG_DIFF"):
            any_diff = True

        rows.append({
            "line_key": line_key,
            "cid_key": cid_key,
            "pattern_segment_tag": p_seg["tag"],
            "test_segment_tag": t_seg["tag"],
            "element_index_raw": i,   # 0 = tag, 1.. = data elements
            "element_position": i,
            "pattern_value": p_val,
            "test_value": t_val,
            "status": status,
            "ignored": bool(ignored),
            "status_effective": status_effective,
            "issue": "",
            "comments": ""
        })

    return rows, any_diff


# -------------------------------------------------------
#  Region compare with N1 alignment (N101 key)
# -------------------------------------------------------
def compare_region_with_n1_key(p_region: List[Dict[str, Any]],
                               t_region: List[Dict[str, Any]],
                               ignore_rules: List[IgnoreRule],
                               seg_diff_rows: List[Dict[str, Any]],
                               elem_diff_rows: List[Dict[str, Any]],
                               missing_rows: List[Dict[str, Any]],
                               extra_rows: List[Dict[str, Any]],
                               line_key: str = "", cid_key: str = ""):
    """
    Compare a small region (e.g., transaction header or line header) but align N1 segments by N101.
    Everything that's not N1 is compared positionally.
    """
    # Split region into N1s and non-N1s, preserving order
    p_n1, p_other = [], []
    for s in p_region:
        (p_n1 if s["tag"] == "N1" else p_other).append(s)
    t_n1, t_other = [], []
    for s in t_region:
        (t_n1 if s["tag"] == "N1" else t_other).append(s)

    # Align N1s by N101
    def n1_key(seg):
        e = seg["elements"]
        return e[1] if len(e) > 1 else ""  # N101

    p_map = defaultdict(list)
    t_map = defaultdict(list)
    p_order, t_order = [], []

    for s in p_n1:
        k = n1_key(s)
        p_map[k].append(s)
        if k not in p_order:
            p_order.append(k)
    for s in t_n1:
        k = n1_key(s)
        t_map[k].append(s)
        if k not in t_order:
            t_order.append(k)

    ordered_keys = list(dict.fromkeys([*p_order, *t_order]))

    for k in ordered_keys:
        pl = p_map.get(k, [])
        tl = t_map.get(k, [])
        c = min(len(pl), len(tl))
        # Matched N1s: element-by-element compare
        for i in range(c):
            rows, _ = compare_elements(pl[i], tl[i], ignore_rules, line_key=line_key, cid_key=cid_key)
            elem_diff_rows.extend(rows)
        # Pattern-only N1s → Missing
        for i in range(c, len(pl)):
            s = pl[i]
            missing_rows.append({
                "pattern_index": s["index"],
                "segment_tag": s["tag"],
                "segment_text": "*".join(s["elements"])
            })
            seg_diff_rows.append({
                "op": "DELETE",
                "meaning": op_meaning("DELETE"),
                "pattern_range": f"{s['index']}:{s['index']+1}",
                "test_range": f"{s['index']}:{s['index']}",
                "pattern_tag": s["tag"],
                "test_tag": ""
            })
        # Test-only N1s → Extra
        for i in range(c, len(tl)):
            s = tl[i]
            extra_rows.append({
                "test_index": s["index"],
                "segment_tag": s["tag"],
                "segment_text": "*".join(s["elements"])
            })
            seg_diff_rows.append({
                "op": "INSERT",
                "meaning": op_meaning("INSERT"),
                "pattern_range": f"{s['index']}:{s['index']}",
                "test_range": f"{s['index']}:{s['index']+1}",
                "pattern_tag": "",
                "test_tag": s["tag"]
            })

    # Compare the non-N1s positionally (unchanged behavior)
    if p_other or t_other:
        positional_diff(p_other, t_other, ignore_rules,
                        seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                        line_key=line_key, cid_key=cid_key)


# -------------------------------------------------------
#  Utilities
# -------------------------------------------------------
def op_meaning(op: str) -> str:
    return {
        "INSERT": "Extra in Test (only in test file)",
        "DELETE": "Missing in Test (present in pattern, missing in test file)",
        "REPLACE": "Values differ in same position",
        "REPLACE_BLOCK": "Block difference (loops or order changed)",
        "EQUAL": "No difference"
    }.get(op, op)


def flatten_elements_for_sheet(parsed: List[Dict[str, Any]], label: str):
    rows = []
    for rec in parsed:
        tag = rec["tag"]; elems = rec["elements"]
        for pos in range(1, len(elems)):
            rows.append({
                "file": label,
                "segment_index": rec["index"],
                "segment_tag": tag,
                "element_position": pos,
                "value": elems[pos],
            })
    return rows


# -------------------------------------------------------
#  856 comparator (per-transaction scope + HL semantic)
# -------------------------------------------------------
def key_for_segment_856(seg: Dict[str, Any]) -> Tuple:
    tag = seg["tag"]
    e = seg["elements"]
    if tag == "REF":
        return ("REF", (e[1] if len(e) > 1 else "") or "")
    if tag == "DTM":
        return ("DTM", (e[1] if len(e) > 1 else "") or "")
    if tag == "N1":
        return ("N1",  (e[1] if len(e) > 1 else "") or "")
    if tag == "MEA":
        return ("MEA",
                (e[1] if len(e) > 1 else "") or "",
                (e[2] if len(e) > 2 else "") or "")
    if tag == "TD1":
        return ("TD1",)
    if tag == "CLD":
        return ("CLD",)
    if tag == "TD5":
        return ("TD5", (e[1] if len(e) > 1 else "") or "")
    if tag == "PRF":
        return ("PRF", "__ANY__")
    if tag == "PID":
        return ("PID", tuple(e[1:]))
    if tag == "LIN":
        return ("LIN",)
    return (tag,)


def get_st_blocks(parsed: List[Dict[str, Any]]) -> List[Tuple[int, int, str, List[Dict[str, Any]]]]:
    blocks = []
    st_positions = [i for i, s in enumerate(parsed) if s["tag"] == "ST"]
    for si, s_idx in enumerate(st_positions):
        e_idx = None
        for j in range(s_idx + 1, len(parsed)):
            if parsed[j]["tag"] == "SE":
                e_idx = j
                break
        end = (e_idx + 1) if e_idx is not None else len(parsed)
        st = parsed[s_idx]
        st_control = st["elements"][2] if len(st["elements"]) > 2 else f"__pos__{len(blocks)}"
        sub = parsed[s_idx:end]
        blocks.append((s_idx, end, st_control, sub))
    return blocks


def header_envelope(parsed: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    first_st = next((i for i, s in enumerate(parsed) if s["tag"] == "ST"), None)
    return parsed[0:first_st] if first_st is not None else parsed[:]


def trailer_envelope(parsed: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    last_se = None
    for i, s in enumerate(parsed):
        if s["tag"] == "SE":
            last_se = i
    return parsed[last_se + 1:] if last_se is not None else []


def transaction_header_region(block: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    first_hl = next((i for i, s in enumerate(block) if s["tag"] == "HL"), None)
    return block[0:first_hl] if first_hl is not None else block[:]


def transaction_header_region_until(block: List[Dict[str, Any]],
                                    start_tag: str) -> List[Dict[str, Any]]:
    first = next((i for i, s in enumerate(block) if s["tag"] == start_tag), None)
    return block[0:first] if first is not None else block[:]


def transaction_trailer_region(block: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    start = None
    for i, s in enumerate(block):
        if s["tag"] == "CTT":
            start = i
            break
    if start is None:
        for i, s in enumerate(block):
            if s["tag"] == "SE":
                start = i
                break
    return block[start:] if start is not None else []


def transaction_hl_region(block: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    hdr = transaction_header_region(block)
    trl = transaction_trailer_region(block)
    start = len(hdr)
    end = len(block) - len(trl) if trl else len(block)
    return block[start:end]


def positional_diff(p_parsed, t_parsed,
                    ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                    line_key: str = "", cid_key: str = ""):
    p_tags = [p["tag"] for p in p_parsed]
    t_tags = [t["tag"] for t in t_parsed]
    opcodes = SequenceMatcher(a=p_tags, b=t_tags, autojunk=False).get_opcodes()

    for tag, i1, i2, j1, j2 in opcodes:
        if tag == "equal":
            for k in range(i2 - i1):
                p_seg = p_parsed[i1 + k]; t_seg = t_parsed[j1 + k]
                rows, any_diff = compare_elements(p_seg, t_seg, ignore_rules, line_key, cid_key)
                elem_diff_rows.extend(rows)
                seg_diff_rows.append({
                    "op": "REPLACE" if any_diff else "EQUAL",
                    "meaning": "Values differ in same position" if any_diff else "No difference",
                    "pattern_range": f"{p_seg['index']}:{p_seg['index']+1}",
                    "test_range": f"{t_seg['index']}:{t_seg['index']+1}",
                    "pattern_tag": p_seg["tag"],
                    "test_tag": t_seg["tag"],
                })
        elif tag == "replace":
            len_p = i2 - i1; len_t = j2 - j1
            if len_p == len_t:
                for k in range(len_p):
                    p_seg = p_parsed[i1 + k]; t_seg = t_parsed[j1 + k]
                    rows, any_diff = compare_elements(p_seg, t_seg, ignore_rules, line_key, cid_key)
                    elem_diff_rows.extend(rows)
                    seg_diff_rows.append({
                        "op": "REPLACE" if any_diff else "EQUAL",
                        "meaning": "Values differ in same position" if any_diff else "No difference",
                        "pattern_range": f"{p_seg['index']}:{p_seg['index']+1}",
                        "test_range": f"{t_seg['index']}:{t_seg['index']+1}",
                        "pattern_tag": p_seg["tag"],
                        "test_tag": t_seg["tag"],
                    })
            else:
                seg_diff_rows.append({
                    "op": "REPLACE_BLOCK",
                    "meaning": "Block difference (loops or order changed)",
                    "pattern_range": f"{i1}:{i2}",
                    "test_range": f"{j1}:{j2}",
                    "pattern_tag": ",".join(p_tags[i1:i2]),
                    "test_tag": ",".join(t_tags[j1:j2]),
                })
        elif tag == "delete":
            for k in range(i1, i2):
                p_seg = p_parsed[k]
                missing_rows.append({
                    "pattern_index": p_seg["index"],
                    "segment_tag": p_seg["tag"],
                    "segment_text": "*".join(p_seg["elements"])
                })
                seg_diff_rows.append({
                    "op": "DELETE",
                    "meaning": "Missing in Test (present in pattern, missing in test file)",
                    "pattern_range": f"{p_seg['index']}:{p_seg['index']+1}",
                    "test_range": f"{t_parsed[j1]['index']}:{t_parsed[j1]['index']}" if (t_parsed and j1 < len(t_parsed)) else f"{p_seg['index']}:{p_seg['index']}",
                    "pattern_tag": p_seg["tag"],
                    "test_tag": ""
                })
        elif tag == "insert":
            for k in range(j1, j2):
                t_seg = t_parsed[k]
                extra_rows.append({
                    "test_index": t_seg["index"],
                    "segment_tag": t_seg["tag"],
                    "segment_text": "*".join(t_seg["elements"])
                })
                seg_diff_rows.append({
                    "op": "INSERT",
                    "meaning": "Extra in Test (only in test file)",
                    "pattern_range": f"{p_parsed[i1]['index']}:{p_parsed[i1]['index']}" if (p_parsed and i1 < len(p_parsed)) else f"{t_seg['index']}:{t_seg['index']}",
                    "test_range": f"{t_seg['index']}:{t_seg['index']+1}",
                    "pattern_tag": "",
                    "test_tag": t_seg["tag"]
                })


def get_hl_blocks(parsed: List[Dict[str, Any]]) -> List[Tuple[int, int]]:
    hl_starts = [i for i, seg in enumerate(parsed) if seg["tag"] == "HL"]
    if not hl_starts:
        return []
    blocks = []
    for s_i, start in enumerate(hl_starts):
        end = len(parsed)
        if s_i + 1 < len(hl_starts):
            end = hl_starts[s_i + 1]
        else:
            for j in range(start + 1, len(parsed)):
                if parsed[j]["tag"] in ("CTT", "SE"):
                    end = j
                    break
        blocks.append((start, end))
    return blocks


def align_hl_block_856(p_block: List[Dict[str, Any]],
                       t_block: List[Dict[str, Any]],
                       ignore_rules: List[IgnoreRule],
                       seg_diff_rows: List[Dict[str, Any]],
                       elem_diff_rows: List[Dict[str, Any]],
                       missing_rows: List[Dict[str, Any]],
                       extra_rows: List[Dict[str, Any]],
                       line_key: str = "", cid_key: str = "") -> None:
    if not p_block or not t_block:
        return

    p_hl = p_block[0]; t_hl = t_block[0]
    rows, any_diff = compare_elements(p_hl, t_hl, ignore_rules, line_key, cid_key)
    elem_diff_rows.extend(rows)
    seg_diff_rows.append({
        "op": "REPLACE" if any_diff else "EQUAL",
        "meaning": "Values differ in same position" if any_diff else "No difference",
        "pattern_range": f"{p_hl['index']}:{p_hl['index']+1}",
        "test_range": f"{t_hl['index']}:{t_hl['index']+1}",
        "pattern_tag": p_hl["tag"],
        "test_tag": t_hl["tag"],
    })

    def build_key_map(block: List[Dict[str, Any]]):
        km = defaultdict(list)
        order = []
        for seg in block[1:]:
            k = key_for_segment_856(seg)
            km[k].append(seg)
            if k not in order:
                order.append(k)
        return km, order

    p_map, p_order = build_key_map(p_block)
    t_map, t_order = build_key_map(t_block)

    seen = set()
    ordered_keys = []
    for k in p_order:
        if k not in seen:
            ordered_keys.append(k); seen.add(k)
    for k in t_order:
        if k not in seen:
            ordered_keys.append(k); seen.add(k)

    for k in ordered_keys:
        p_list = p_map.get(k, [])
        t_list = t_map.get(k, [])
        n_common = min(len(p_list), len(t_list))

        for i in range(n_common):
            p_seg = p_list[i]; t_seg = t_list[i]
            rows, any_diff = compare_elements(p_seg, t_seg, ignore_rules, line_key, cid_key)
            elem_diff_rows.extend(rows)
            seg_diff_rows.append({
                "op": "REPLACE" if any_diff else "EQUAL",
                "meaning": "Values differ in same position" if any_diff else "No difference",
                "pattern_range": f"{p_seg['index']}:{p_seg['index']+1}",
                "test_range": f"{t_seg['index']}:{t_seg['index']+1}",
                "pattern_tag": p_seg["tag"],
                "test_tag": t_seg["tag"],
            })

        for i in range(n_common, len(p_list)):
            p_seg = p_list[i]
            missing_rows.append({
                "pattern_index": p_seg["index"],
                "segment_tag": p_seg["tag"],
                "segment_text": "*".join(p_seg["elements"])
            })
            seg_diff_rows.append({
                "op": "DELETE",
                "meaning": "Missing in Test (present in pattern, missing in test file)",
                "pattern_range": f"{p_seg['index']}:{p_seg['index']+1}",
                "test_range": f"{t_block[-1]['index']+1}:{t_block[-1]['index']+1}",
                "pattern_tag": p_seg["tag"],
                "test_tag": ""
            })

        for i in range(n_common, len(t_list)):
            t_seg = t_list[i]
            extra_rows.append({
                "test_index": t_seg["index"],
                "segment_tag": t_seg["tag"],
                "segment_text": "*".join(t_seg["elements"])
            })
            seg_diff_rows.append({
                "op": "INSERT",
                "meaning": op_meaning("INSERT"),
                "pattern_range": f"{p_block[-1]['index']+1}:{p_block[-1]['index']+1}",
                "test_range": f"{t_seg['index']}:{t_seg['index']+1}",
                "pattern_tag": "",
                "test_tag": t_seg["tag"]
            })


def compare_856_transactions(p_parsed, t_parsed, ignore_rules,
                             seg_diff_rows, elem_diff_rows, missing_rows, extra_rows):
    # Envelope header
    p_env_hdr = header_envelope(p_parsed)
    t_env_hdr = header_envelope(t_parsed)
    positional_diff(p_env_hdr, t_env_hdr, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

    # ST blocks paired by ST02
    p_tx = get_st_blocks(p_parsed)
    t_tx = get_st_blocks(t_parsed)
    p_map = {ctrl: blk for (_, _, ctrl, blk) in p_tx}
    t_map = {ctrl: blk for (_, _, ctrl, blk) in t_tx}
    keys_in_order = [ctrl for (_, _, ctrl, _) in p_tx]
    for (_, _, ctrl, _) in t_tx:
        if ctrl not in keys_in_order:
            keys_in_order.append(ctrl)

    for ctrl in keys_in_order:
        p_blk = p_map.get(ctrl)
        t_blk = t_map.get(ctrl)

        if p_blk and t_blk:
            # Transaction header (pre-HL) — N1-aligned
            p_hdr = transaction_header_region(p_blk)
            t_hdr = transaction_header_region(t_blk)
            compare_region_with_n1_key(p_hdr, t_hdr, ignore_rules,
                                       seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                                       line_key="", cid_key="")

            # HL regions
            p_hl_region = transaction_hl_region(p_blk)
            t_hl_region = transaction_hl_region(t_blk)

            p_hl_blocks_idx = get_hl_blocks(p_hl_region)
            t_hl_blocks_idx = get_hl_blocks(t_hl_region)

            p_hl_blocks = [(p_hl_region[s:e]) for (s, e) in p_hl_blocks_idx]
            t_hl_blocks = [(t_hl_region[s:e]) for (s, e) in t_hl_blocks_idx]

            n_common = min(len(p_hl_blocks), len(t_hl_blocks))
            for i in range(n_common):
                align_hl_block_856(p_hl_blocks[i], t_hl_blocks[i], ignore_rules,
                                   seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

            for i in range(n_common, len(p_hl_blocks)):
                for seg in p_hl_blocks[i]:
                    if seg["tag"] == "HL":
                        seg_diff_rows.append({
                            "op": "DELETE",
                            "meaning": op_meaning("DELETE"),
                            "pattern_range": f"{seg['index']}:{seg['index']+1}",
                            "test_range": f"{seg['index']}:{seg['index']}",
                            "pattern_tag": seg["tag"],
                            "test_tag": ""
                        })
                    missing_rows.append({
                        "pattern_index": seg["index"],
                        "segment_tag": seg["tag"],
                        "segment_text": "*".join(seg["elements"])
                    })

            for i in range(n_common, len(t_hl_blocks)):
                for seg in t_hl_blocks[i]:
                    if seg["tag"] == "HL":
                        seg_diff_rows.append({
                            "op": "INSERT",
                            "meaning": op_meaning("INSERT"),
                            "pattern_range": f"{seg['index']}:{seg['index']}",
                            "test_range": f"{seg['index']}:{seg['index']+1}",
                            "pattern_tag": "",
                            "test_tag": seg["tag"]
                        })
                    extra_rows.append({
                        "test_index": seg["index"],
                        "segment_tag": seg["tag"],
                        "segment_text": "*".join(seg["elements"])
                    })

            # Trailer (CTT/SE)
            p_trl = transaction_trailer_region(p_blk)
            t_trl = transaction_trailer_region(t_blk)
            positional_diff(p_trl, t_trl, ignore_rules,
                            seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

        elif p_blk and not t_blk:
            for seg in p_blk:
                if seg["tag"] == "ST":
                    seg_diff_rows.append({
                        "op": "DELETE",
                        "meaning": op_meaning("DELETE"),
                        "pattern_range": f"{seg['index']}:{seg['index']+1}",
                        "test_range": f"{seg['index']}:{seg['index']}",
                        "pattern_tag": seg["tag"],
                        "test_tag": ""
                    })
                missing_rows.append({
                    "pattern_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })
        elif t_blk and not p_blk:
            for seg in t_blk:
                if seg["tag"] == "ST":
                    seg_diff_rows.append({
                        "op": "INSERT",
                        "meaning": op_meaning("INSERT"),
                        "pattern_range": f"{seg['index']}:{seg['index']}",
                        "test_range": f"{seg['index']}:{seg['index']+1}",
                        "pattern_tag": "",
                        "test_tag": seg["tag"]
                    })
                extra_rows.append({
                    "test_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })

    # Envelope trailer (GE/IEA)
    p_env_trl = trailer_envelope(p_parsed)
    t_env_trl = trailer_envelope(t_parsed)
    positional_diff(p_env_trl, t_env_trl, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)


# -------------------------------------------------------
#  863 comparator (ST/SE → LIN lines → CID groups keyed by MEA01/MEA02)
# -------------------------------------------------------
LIN_ID_PRIORITY = ["BP", "PO", "VN", "SN", "HN", "TU"]


def lin_line_key(elems: List[str]) -> str:
    qual_to_val = {}
    i = 2
    while i + 1 < len(elems):
        qual = elems[i]
        val = elems[i + 1]
        if qual:
            qual_to_val[qual] = val
        i += 2
    for qual in LIN_ID_PRIORITY:
        if qual in qual_to_val and qual_to_val[qual]:
            return f"{qual}:{qual_to_val[qual]}"
    if qual_to_val:
        return "|".join(f"{k}:{v}" for k, v in qual_to_val.items())
    return elems[1] if len(elems) > 1 else "__LIN__"


def get_lin_blocks(tx_block: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    starts = [i for i, s in enumerate(tx_block) if s["tag"] == "LIN"]
    if not starts:
        return []
    blocks = []
    for si, start in enumerate(starts):
        end = len(tx_block)
        if si + 1 < len(starts):
            end = starts[si + 1]
        else:
            for j in range(start + 1, len(tx_block)):
                if tx_block[j]["tag"] in ("CTT", "SE"):
                    end = j
                    break
        blocks.append(tx_block[start:end])
    return blocks


def split_cid_groups(line_block: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    groups = []
    current = []
    for seg in line_block:
        if seg["tag"] == "CID":
            if current:
                groups.append(current)
            current = [seg]
        else:
            if current:
                if seg["tag"] in ("LIN", "CTT", "SE"):
                    groups.append(current)
                    current = []
                    continue
                current.append(seg)
    if current:
        groups.append(current)
    return groups


def cid_key_from_group(cid_group: List[Dict[str, Any]]) -> str:
    """Return a stable identifier for all CID groups that share the same CID qualifier."""
    cid_seg = next((s for s in cid_group if s["tag"] == "CID"), None)
    if not cid_seg:
        return "__NO_CID__"
    elems = cid_seg["elements"]
    cid01 = elems[1] if len(elems) > 1 else ""
    cid02 = elems[2] if len(elems) > 2 else ""
    cid03 = elems[3] if len(elems) > 3 else ""
    return f"{cid01}:{cid02}:{cid03}"


def mea_signature(seg: Dict[str, Any]) -> Tuple[str, str, str]:
    elems = seg["elements"]
    return (
        elems[1] if len(elems) > 1 else "",
        elems[2] if len(elems) > 2 else "",
        elems[4] if len(elems) > 4 else ""
    )


def cid_group_sort_key(cid_group: List[Dict[str, Any]]) -> Tuple[Tuple[str, str, str], ...]:
    mea_keys = [mea_signature(seg) for seg in cid_group if seg["tag"] == "MEA"]
    if not mea_keys:
        return (('__NO_MEA__', '', ''),)
    return tuple(sorted(mea_keys))


def line_header_region(line_block: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    first_cid = next((i for i, s in enumerate(line_block) if s["tag"] == "CID"), None)
    return line_block[:first_cid] if first_cid is not None else line_block[:]


def line_trailer_region(line_block: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Original logic returned everything after the last CID, which duplicates MEAs from the
    final CID group in Element Diff/Missing tabs. There typically are no true "trailer"
    segments for 863 lines once CID groups start, so we return an empty list to avoid
    re-processing those segments.
    """
    return []


def get_rcd_blocks(tx_block: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    starts = [i for i, s in enumerate(tx_block) if s["tag"] == "RCD"]
    if not starts:
        return []
    blocks = []
    for si, start in enumerate(starts):
        end = len(tx_block)
        if si + 1 < len(starts):
            end = starts[si + 1]
        else:
            for j in range(start + 1, len(tx_block)):
                if tx_block[j]["tag"] in ("CTT", "SE"):
                    end = j
                    break
        blocks.append(tx_block[start:end])
    return blocks


def rcd_block_key(block: List[Dict[str, Any]]) -> str:
    lin_seg = next((s for s in block if s["tag"] == "LIN"), None)
    if lin_seg:
        return lin_line_key(lin_seg["elements"])
    rcd_seg = block[0] if block else None
    if not rcd_seg:
        return "__RCD__"
    elems = rcd_seg["elements"]
    if len(elems) > 1:
        return "RCD:" + "|".join(elems[1:])
    return f"RCD_INDEX_{rcd_seg['index']}"


def compare_rcd_block(p_block: List[Dict[str, Any]],
                      t_block: List[Dict[str, Any]],
                      ignore_rules: List[IgnoreRule],
                      seg_diff_rows: List[Dict[str, Any]],
                      elem_diff_rows: List[Dict[str, Any]],
                      missing_rows: List[Dict[str, Any]],
                      extra_rows: List[Dict[str, Any]],
                      line_key: str) -> None:
    if not p_block or not t_block:
        return

    p_rcd = p_block[0]
    t_rcd = t_block[0]
    rows, any_diff = compare_elements(p_rcd, t_rcd, ignore_rules, line_key=line_key, cid_key="")
    elem_diff_rows.extend(rows)
    seg_diff_rows.append({
        "op": "REPLACE" if any_diff else "EQUAL",
        "meaning": op_meaning("REPLACE" if any_diff else "EQUAL"),
        "pattern_range": f"{p_rcd['index']}:{p_rcd['index']+1}",
        "test_range": f"{t_rcd['index']}:{t_rcd['index']+1}",
        "pattern_tag": p_rcd["tag"],
        "test_tag": t_rcd["tag"],
    })

    known_tags = ("LIN", "PID", "REF", "DTM", "MEA")
    p_known = {tag: [] for tag in known_tags}
    t_known = {tag: [] for tag in known_tags}
    p_other: List[Dict[str, Any]] = []
    t_other: List[Dict[str, Any]] = []

    for seg in p_block[1:]:
        if seg["tag"] in p_known:
            p_known[seg["tag"]].append(seg)
        else:
            p_other.append(seg)
    for seg in t_block[1:]:
        if seg["tag"] in t_known:
            t_known[seg["tag"]].append(seg)
        else:
            t_other.append(seg)

    def align_segments(tag: str, key_fn):
        p_list = p_known[tag]
        t_list = t_known[tag]
        if not p_list and not t_list:
            return

        p_map = defaultdict(list)
        t_map = defaultdict(list)
        p_order: List[Any] = []
        t_order: List[Any] = []

        for seg in p_list:
            key = key_fn(seg)
            p_map[key].append(seg)
            if key not in p_order:
                p_order.append(key)

        for seg in t_list:
            key = key_fn(seg)
            t_map[key].append(seg)
            if key not in t_order:
                t_order.append(key)

        ordered_keys: List[Any] = []
        seen = set()
        for key in [*p_order, *t_order]:
            if key not in seen:
                ordered_keys.append(key)
                seen.add(key)

        for key in ordered_keys:
            pl = p_map.get(key, [])
            tl = t_map.get(key, [])
            common = min(len(pl), len(tl))
            for i in range(common):
                p_seg = pl[i]
                t_seg = tl[i]
                rows, seg_diff = compare_elements(p_seg, t_seg, ignore_rules, line_key=line_key, cid_key="")
                elem_diff_rows.extend(rows)
                seg_diff_rows.append({
                    "op": "REPLACE" if seg_diff else "EQUAL",
                    "meaning": op_meaning("REPLACE" if seg_diff else "EQUAL"),
                    "pattern_range": f"{p_seg['index']}:{p_seg['index']+1}",
                    "test_range": f"{t_seg['index']}:{t_seg['index']+1}",
                    "pattern_tag": p_seg["tag"],
                    "test_tag": t_seg["tag"],
                })
            for i in range(common, len(pl)):
                seg = pl[i]
                missing_rows.append({
                    "pattern_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })
                seg_diff_rows.append({
                    "op": "DELETE",
                    "meaning": op_meaning("DELETE"),
                    "pattern_range": f"{seg['index']}:{seg['index']+1}",
                    "test_range": f"{seg['index']}:{seg['index']}",
                    "pattern_tag": seg["tag"],
                    "test_tag": ""
                })
            for i in range(common, len(tl)):
                seg = tl[i]
                extra_rows.append({
                    "test_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })
                seg_diff_rows.append({
                    "op": "INSERT",
                    "meaning": op_meaning("INSERT"),
                    "pattern_range": f"{seg['index']}:{seg['index']}",
                    "test_range": f"{seg['index']}:{seg['index']+1}",
                    "pattern_tag": "",
                    "test_tag": seg["tag"]
                })

    align_segments("LIN", lambda seg: lin_line_key(seg["elements"]))
    align_segments("PID", lambda seg: tuple(seg["elements"][1:]))
    align_segments("REF", lambda seg: seg["elements"][1] if len(seg["elements"]) > 1 else "")
    align_segments("DTM", lambda seg: seg["elements"][1] if len(seg["elements"]) > 1 else "")
    align_segments("MEA", mea_signature)

    if p_other or t_other:
        positional_diff(p_other, t_other, ignore_rules,
                        seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                        line_key=line_key, cid_key="")


def compare_861_transactions(p_parsed, t_parsed, ignore_rules,
                             seg_diff_rows, elem_diff_rows, missing_rows, extra_rows):
    # Envelope header
    p_env_hdr = header_envelope(p_parsed)
    t_env_hdr = header_envelope(t_parsed)
    positional_diff(p_env_hdr, t_env_hdr, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

    # ST blocks paired by ST02
    p_tx = get_st_blocks(p_parsed)
    t_tx = get_st_blocks(t_parsed)
    p_map = {ctrl: blk for (_, _, ctrl, blk) in p_tx}
    t_map = {ctrl: blk for (_, _, ctrl, blk) in t_tx}
    keys_in_order = [ctrl for (_, _, ctrl, _) in p_tx]
    for (_, _, ctrl, _) in t_tx:
        if ctrl not in keys_in_order:
            keys_in_order.append(ctrl)

    for ctrl in keys_in_order:
        p_blk = p_map.get(ctrl)
        t_blk = t_map.get(ctrl)

        if p_blk and t_blk:
            # Header up to first RCD — align N1 segments by qualifier
            p_head = transaction_header_region_until(p_blk, "RCD")
            t_head = transaction_header_region_until(t_blk, "RCD")
            compare_region_with_n1_key(p_head, t_head, ignore_rules,
                                       seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                                       line_key="", cid_key="")

            # Item blocks keyed by LIN identifiers or RCD content
            p_blocks = get_rcd_blocks(p_blk)
            t_blocks = get_rcd_blocks(t_blk)

            p_block_map = defaultdict(list)
            p_block_order: List[str] = []
            for block in p_blocks:
                key = rcd_block_key(block)
                p_block_map[key].append(block)
                if key not in p_block_order:
                    p_block_order.append(key)

            t_block_map = defaultdict(list)
            t_block_order: List[str] = []
            for block in t_blocks:
                key = rcd_block_key(block)
                t_block_map[key].append(block)
                if key not in t_block_order:
                    t_block_order.append(key)

            ordered_keys: List[str] = []
            seen = set()
            for key in [*p_block_order, *t_block_order]:
                if key not in seen:
                    ordered_keys.append(key)
                    seen.add(key)

            for key in ordered_keys:
                p_list = p_block_map.get(key, [])
                t_list = t_block_map.get(key, [])
                common = min(len(p_list), len(t_list))
                for i in range(common):
                    compare_rcd_block(p_list[i], t_list[i], ignore_rules,
                                      seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                                      line_key=key)
                for i in range(common, len(p_list)):
                    for seg in p_list[i]:
                        missing_rows.append({
                            "pattern_index": seg["index"],
                            "segment_tag": seg["tag"],
                            "segment_text": "*".join(seg["elements"])
                        })
                for i in range(common, len(t_list)):
                    for seg in t_list[i]:
                        extra_rows.append({
                            "test_index": seg["index"],
                            "segment_tag": seg["tag"],
                            "segment_text": "*".join(seg["elements"])
                        })

            # Trailer (CTT/SE)
            p_trl = transaction_trailer_region(p_blk)
            t_trl = transaction_trailer_region(t_blk)
            positional_diff(p_trl, t_trl, ignore_rules,
                            seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

        elif p_blk and not t_blk:
            for seg in p_blk:
                if seg["tag"] == "ST":
                    seg_diff_rows.append({
                        "op": "DELETE",
                        "meaning": op_meaning("DELETE"),
                        "pattern_range": f"{seg['index']}:{seg['index']+1}",
                        "test_range": f"{seg['index']}:{seg['index']}",
                        "pattern_tag": seg["tag"],
                        "test_tag": ""
                    })
                missing_rows.append({
                    "pattern_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })
        elif t_blk and not p_blk:
            for seg in t_blk:
                if seg["tag"] == "ST":
                    seg_diff_rows.append({
                        "op": "INSERT",
                        "meaning": op_meaning("INSERT"),
                        "pattern_range": f"{seg['index']}:{seg['index']}",
                        "test_range": f"{seg['index']}:{seg['index']+1}",
                        "pattern_tag": "",
                        "test_tag": seg["tag"]
                    })
                extra_rows.append({
                    "test_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })

    # Envelope trailer (GE/IEA)
    p_env_trl = trailer_envelope(p_parsed)
    t_env_trl = trailer_envelope(t_parsed)
    positional_diff(p_env_trl, t_env_trl, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)


def compare_863_transactions(p_parsed, t_parsed, ignore_rules,
                             seg_diff_rows, elem_diff_rows, missing_rows, extra_rows):
    # Envelope header
    p_env_hdr = header_envelope(p_parsed)
    t_env_hdr = header_envelope(t_parsed)
    positional_diff(p_env_hdr, t_env_hdr, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

    # ST blocks paired by ST02
    p_tx = get_st_blocks(p_parsed)
    t_tx = get_st_blocks(t_parsed)
    p_map = {ctrl: blk for (_, _, ctrl, blk) in p_tx}
    t_map = {ctrl: blk for (_, _, ctrl, blk) in t_tx}
    keys_in_order = [ctrl for (_, _, ctrl, _) in p_tx]
    for (_, _, ctrl, _) in t_tx:
        if ctrl not in keys_in_order:
            keys_in_order.append(ctrl)

    for ctrl in keys_in_order:
        p_blk = p_map.get(ctrl)
        t_blk = t_map.get(ctrl)

        if p_blk and t_blk:
            # Transaction header before first LIN — N1-aligned
            p_head = []
            for rec in p_blk:
                if rec["tag"] == "LIN": break
                p_head.append(rec)
            t_head = []
            for rec in t_blk:
                if rec["tag"] == "LIN": break
                t_head.append(rec)
            compare_region_with_n1_key(p_head, t_head, ignore_rules,
                                       seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                                       line_key="", cid_key="")

            # Line blocks
            p_lines = get_lin_blocks(p_blk)
            t_lines = get_lin_blocks(t_blk)

            def line_key_for_block(line_block):
                lin = line_block[0]  # first is LIN
                return lin_line_key(lin["elements"])

            p_line_map = {}
            p_line_order = []
            for lb in p_lines:
                k = line_key_for_block(lb)
                p_line_map.setdefault(k, []).append(lb)
                if k not in p_line_order:
                    p_line_order.append(k)

            t_line_map = {}
            t_line_order = []
            for lb in t_lines:
                k = line_key_for_block(lb)
                t_line_map.setdefault(k, []).append(lb)
                if k not in t_line_order:
                    t_line_order.append(k)

            ordered_line_keys = []
            seen = set()
            for k in p_line_order:
                if k not in seen:
                    ordered_line_keys.append(k); seen.add(k)
            for k in t_line_order:
                if k not in seen:
                    ordered_line_keys.append(k); seen.add(k)

            for lk in ordered_line_keys:
                p_lbs = p_line_map.get(lk, [])
                t_lbs = t_line_map.get(lk, [])
                n_common = min(len(p_lbs), len(t_lbs))

                for i in range(n_common):
                    p_line = p_lbs[i]
                    t_line = t_lbs[i]

                    # Compare line header (LIN + DTM/REF before first CID) — N1-aligned
                    p_line_head = line_header_region(p_line)
                    t_line_head = line_header_region(t_line)
                    compare_region_with_n1_key(p_line_head, t_line_head, ignore_rules,
                                               seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                                               line_key=lk, cid_key="")

                    # Split into CID groups and compare keyed by MEA01:MEA02
                    p_cids = split_cid_groups(p_line)
                    t_cids = split_cid_groups(t_line)

                    def map_cids(cid_groups):
                        cmap = defaultdict(list)
                        order = []
                        for g in cid_groups:
                            key = cid_key_from_group(g)
                            cmap[key].append(g)
                            if key not in order:
                                order.append(key)
                        return cmap, order

                    p_cmap, p_ckeys = map_cids(p_cids)
                    t_cmap, t_ckeys = map_cids(t_cids)

                    ordered_cid_keys = []
                    cseen = set()
                    for ck in p_ckeys:
                        if ck not in cseen:
                            ordered_cid_keys.append(ck); cseen.add(ck)
                    for ck in t_ckeys:
                        if ck not in cseen:
                            ordered_cid_keys.append(ck); cseen.add(ck)

                    for ck in ordered_cid_keys:
                        p_groups = sorted(p_cmap.get(ck, []), key=cid_group_sort_key)
                        t_groups = sorted(t_cmap.get(ck, []), key=cid_group_sort_key)
                        m = min(len(p_groups), len(t_groups))

                        for gi in range(m):
                            pg = p_groups[gi]
                            tg = t_groups[gi]

                            # Compare CID heads (CID vs CID)
                            p_cid = next((s for s in pg if s["tag"] == "CID"), None)
                            t_cid = next((s for s in tg if s["tag"] == "CID"), None)
                            if p_cid and t_cid:
                                rows, _ = compare_elements(p_cid, t_cid, ignore_rules, line_key=lk, cid_key=ck)
                                elem_diff_rows.extend(rows)

                            # TMD required logic
                            p_tmd = [s for s in pg if s["tag"] == "TMD"]
                            t_tmd = [s for s in tg if s["tag"] == "TMD"]
                            if p_tmd and not t_tmd:
                                for s in p_tmd:
                                    missing_rows.append({
                                        "pattern_index": s["index"],
                                        "segment_tag": s["tag"],
                                        "segment_text": "*".join(s["elements"])
                                    })
                                    seg_diff_rows.append({
                                        "op": "DELETE",
                                        "meaning": op_meaning("DELETE"),
                                        "pattern_range": f"{s['index']}:{s['index']+1}",
                                        "test_range": f"{s['index']}:{s['index']}",
                                        "pattern_tag": s["tag"],
                                        "test_tag": ""
                                    })
                            elif t_tmd and not p_tmd:
                                for s in t_tmd:
                                    extra_rows.append({
                                        "test_index": s["index"],
                                        "segment_tag": s["tag"],
                                        "segment_text": "*".join(s["elements"])
                                    })
                                    seg_diff_rows.append({
                                        "op": "INSERT",
                                        "meaning": op_meaning("INSERT"),
                                        "pattern_range": f"{s['index']}:{s['index']}",
                                        "test_range": f"{s['index']}:{s['index']+1}",
                                        "pattern_tag": "",
                                        "test_tag": s["tag"]
                                    })
                            else:
                                c = min(len(p_tmd), len(t_tmd))
                                for ti in range(c):
                                    rows, _ = compare_elements(p_tmd[ti], t_tmd[ti], ignore_rules, line_key=lk, cid_key=ck)
                                    elem_diff_rows.extend(rows)
                                for ti in range(c, len(p_tmd)):
                                    s = p_tmd[ti]
                                    missing_rows.append({
                                        "pattern_index": s["index"],
                                        "segment_tag": s["tag"],
                                        "segment_text": "*".join(s["elements"])
                                    })
                                for ti in range(c, len(t_tmd)):
                                    s = t_tmd[ti]
                                    extra_rows.append({
                                        "test_index": s["index"],
                                        "segment_tag": s["tag"],
                                        "segment_text": "*".join(s["elements"])
                                    })

                            # MEA alignment by (MEA01,MEA02,MEA04)
                            def mea_key(seg):
                                return mea_signature(seg)

                            p_mea_map = defaultdict(list)
                            t_mea_map = defaultdict(list)
                            for s in pg:
                                if s["tag"] == "MEA":
                                    p_mea_map[mea_key(s)].append(s)
                            for s in tg:
                                if s["tag"] == "MEA":
                                    t_mea_map[mea_key(s)].append(s)

                            all_mea_keys = sorted(set(p_mea_map.keys()) | set(t_mea_map.keys()))

                            for mk in all_mea_keys:
                                pml = p_mea_map.get(mk, [])
                                tml = t_mea_map.get(mk, [])
                                c2 = min(len(pml), len(tml))
                                for ii in range(c2):
                                    rows, _ = compare_elements(pml[ii], tml[ii], ignore_rules, line_key=lk, cid_key=ck)
                                    elem_diff_rows.extend(rows)
                                for ii in range(c2, len(pml)):
                                    s = pml[ii]
                                    missing_rows.append({
                                        "pattern_index": s["index"],
                                        "segment_tag": s["tag"],
                                        "segment_text": "*".join(s["elements"])
                                    })
                                for ii in range(c2, len(tml)):
                                    s = tml[ii]
                                    extra_rows.append({
                                        "test_index": s["index"],
                                        "segment_tag": s["tag"],
                                        "segment_text": "*".join(s["elements"])
                                    })

                        # Unpaired groups
                        for gi in range(m, len(p_groups)):
                            for s in p_groups[gi]:
                                missing_rows.append({
                                    "pattern_index": s["index"],
                                    "segment_tag": s["tag"],
                                    "segment_text": "*".join(s["elements"])
                                })
                        for gi in range(m, len(t_groups)):
                            for s in t_groups[gi]:
                                extra_rows.append({
                                    "test_index": s["index"],
                                    "segment_tag": s["tag"],
                                    "segment_text": "*".join(s["elements"])
                                })

                    # Compare any line trailer (after last CID)
                    p_line_tail = line_trailer_region(p_line)
                    t_line_tail = line_trailer_region(t_line)
                    if p_line_tail or t_line_tail:
                        positional_diff(p_line_tail, t_line_tail, ignore_rules,
                                        seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                                        line_key=lk, cid_key="")

                # Pattern-only lines
                for i in range(n_common, len(p_lbs)):
                    for s in p_lbs[i]:
                        missing_rows.append({
                            "pattern_index": s["index"],
                            "segment_tag": s["tag"],
                            "segment_text": "*".join(s["elements"])
                        })
                # Test-only lines
                for i in range(n_common, len(t_lbs)):
                    for s in t_lbs[i]:
                        extra_rows.append({
                            "test_index": s["index"],
                            "segment_tag": s["tag"],
                            "segment_text": "*".join(s["elements"])
                        })

            # Trailer (CTT/SE)
            p_trl = transaction_trailer_region(p_blk)
            t_trl = transaction_trailer_region(t_blk)
            positional_diff(p_trl, t_trl, ignore_rules,
                            seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

        elif p_blk and not t_blk:
            for seg in p_blk:
                if seg["tag"] == "ST":
                    seg_diff_rows.append({
                        "op": "DELETE",
                        "meaning": op_meaning("DELETE"),
                        "pattern_range": f"{seg['index']}:{seg['index']+1}",
                        "test_range": f"{seg['index']}:{seg['index']}",
                        "pattern_tag": seg["tag"],
                        "test_tag": ""
                    })
                missing_rows.append({
                    "pattern_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })
        elif t_blk and not p_blk:
            for seg in t_blk:
                if seg["tag"] == "ST":
                    seg_diff_rows.append({
                        "op": "INSERT",
                        "meaning": op_meaning("INSERT"),
                        "pattern_range": f"{seg['index']}:{seg['index']}",
                        "test_range": f"{seg['index']}:{seg['index']+1}",
                        "pattern_tag": "",
                        "test_tag": seg["tag"]
                    })
                extra_rows.append({
                    "test_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })

    # Envelope trailer (GE/IEA)
    p_env_trl = trailer_envelope(p_parsed)
    t_env_trl = trailer_envelope(t_parsed)
    positional_diff(p_env_trl, t_env_trl, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)


# -------------------------------------------------------
#  EDIFACT comparator (UNH/UNT aligned)
# -------------------------------------------------------
def edifact_header_envelope(parsed: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    first_unh = next((i for i, s in enumerate(parsed) if s["tag"] == "UNH"), None)
    return parsed[0:first_unh] if first_unh is not None else parsed[:]


def edifact_trailer_envelope(parsed: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    last_unt = None
    for i, s in enumerate(parsed):
        if s["tag"] == "UNT":
            last_unt = i
    return parsed[last_unt + 1:] if last_unt is not None else []


def get_unh_blocks(parsed: List[Dict[str, Any]]) -> List[Tuple[int, int, str, List[Dict[str, Any]]]]:
    blocks = []
    for idx, seg in enumerate(parsed):
        if seg["tag"] != "UNH":
            continue
        end = None
        for j in range(idx + 1, len(parsed)):
            if parsed[j]["tag"] == "UNT":
                end = j + 1
                break
        if end is None:
            end = len(parsed)
        ref = seg["elements"][1] if len(seg["elements"]) > 1 else f"__pos__{len(blocks)}"
        blocks.append((idx, end, ref, parsed[idx:end]))
    return blocks


def compare_edifact_transactions(tx_type: str,
                                 p_parsed: List[Dict[str, Any]],
                                 t_parsed: List[Dict[str, Any]],
                                 ignore_rules: List[IgnoreRule],
                                 seg_diff_rows: List[Dict[str, Any]],
                                 elem_diff_rows: List[Dict[str, Any]],
                                 missing_rows: List[Dict[str, Any]],
                                 extra_rows: List[Dict[str, Any]]) -> None:
    p_env_hdr = edifact_header_envelope(p_parsed)
    t_env_hdr = edifact_header_envelope(t_parsed)
    positional_diff(p_env_hdr, t_env_hdr, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)

    p_blocks = get_unh_blocks(p_parsed)
    t_blocks = get_unh_blocks(t_parsed)

    p_map = {ref: blk for (_, _, ref, blk) in p_blocks}
    t_map = {ref: blk for (_, _, ref, blk) in t_blocks}

    keys_in_order: List[str] = []
    for (_, _, ref, _) in p_blocks:
        if ref not in keys_in_order:
            keys_in_order.append(ref)
    for (_, _, ref, _) in t_blocks:
        if ref not in keys_in_order:
            keys_in_order.append(ref)

    for ref in keys_in_order:
        p_blk = p_map.get(ref)
        t_blk = t_map.get(ref)
        if p_blk and t_blk:
            positional_diff(p_blk, t_blk, ignore_rules,
                            seg_diff_rows, elem_diff_rows, missing_rows, extra_rows,
                            line_key=ref, cid_key="")
        elif p_blk and not t_blk:
            for seg in p_blk:
                if seg["tag"] == "UNH":
                    seg_diff_rows.append({
                        "op": "DELETE",
                        "meaning": op_meaning("DELETE"),
                        "pattern_range": f"{seg['index']}:{seg['index']+1}",
                        "test_range": f"{seg['index']}:{seg['index']}",
                        "pattern_tag": seg["tag"],
                        "test_tag": ""
                    })
                missing_rows.append({
                    "pattern_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })
        elif t_blk and not p_blk:
            for seg in t_blk:
                if seg["tag"] == "UNH":
                    seg_diff_rows.append({
                        "op": "INSERT",
                        "meaning": op_meaning("INSERT"),
                        "pattern_range": f"{seg['index']}:{seg['index']}",
                        "test_range": f"{seg['index']}:{seg['index']+1}",
                        "pattern_tag": "",
                        "test_tag": seg["tag"]
                    })
                extra_rows.append({
                    "test_index": seg["index"],
                    "segment_tag": seg["tag"],
                    "segment_text": "*".join(seg["elements"])
                })

    p_env_trl = edifact_trailer_envelope(p_parsed)
    t_env_trl = edifact_trailer_envelope(t_parsed)
    positional_diff(p_env_trl, t_env_trl, ignore_rules,
                    seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)


# -------------------------------------------------------
#  Report builder
# -------------------------------------------------------
def extract_env(parsed: List[Dict[str, Any]], edi_format: str) -> Dict[str, str]:
    d = {}
    tags = {p["tag"]: p for p in parsed}
    if edi_format == EDI_FORMAT_X12:
        isa = tags.get("ISA"); gs = tags.get("GS"); st = tags.get("ST")
        se = tags.get("SE"); ge = tags.get("GE"); iea = tags.get("IEA")
        if isa:
            d["ISA_ControlVersion"] = isa["elements"][12] if len(isa["elements"]) > 12 else ""
            d["ISA_ControlNumber"] = isa["elements"][13] if len(isa["elements"]) > 13 else ""
            d["ISA_UsageIndicator"] = isa["elements"][15] if len(isa["elements"]) > 15 else ""
        if gs:
            d["GS_FunctionalID"] = gs["elements"][1] if len(gs["elements"]) > 1 else ""
            d["GS_Version"] = gs["elements"][8] if len(gs["elements"]) > 8 else ""
            d["GS_GroupControl"] = gs["elements"][6] if len(gs["elements"]) > 6 else ""
        if st:
            d["ST_SetID"] = st["elements"][1] if len(st["elements"]) > 1 else ""
            d["ST_Control"] = st["elements"][2] if len(st["elements"]) > 2 else ""
        if se:
            d["SE_SegmentCount"] = se["elements"][1] if len(se["elements"]) > 1 else ""
            d["SE_Control"] = se["elements"][2] if len(se["elements"]) > 2 else ""
        if ge:
            d["GE_TransactionCount"] = ge["elements"][1] if len(ge["elements"]) > 1 else ""
            d["GE_GroupControl"] = ge["elements"][2] if len(ge["elements"]) > 2 else ""
        if iea:
            d["IEA_GroupCount"] = iea["elements"][1] if len(iea["elements"]) > 1 else ""
            d["IEA_InterchangeControl"] = iea["elements"][2] if len(iea["elements"]) > 2 else ""
    else:
        unb = tags.get("UNB"); unh = tags.get("UNH"); unt = tags.get("UNT"); unz = tags.get("UNZ")
        if unb:
            d["UNB_Syntax"] = unb["elements"][1] if len(unb["elements"]) > 1 else ""
            d["UNB_Sender"] = unb["elements"][2] if len(unb["elements"]) > 2 else ""
            d["UNB_Recipient"] = unb["elements"][3] if len(unb["elements"]) > 3 else ""
            d["UNB_DateTime"] = unb["elements"][4] if len(unb["elements"]) > 4 else ""
            d["UNB_ControlReference"] = unb["elements"][5] if len(unb["elements"]) > 5 else ""
        if unh:
            d["UNH_MessageRef"] = unh["elements"][1] if len(unh["elements"]) > 1 else ""
            if len(unh["elements"]) > 2 and unh["elements"][2]:
                d["UNH_MessageType"] = unh["elements"][2].split(":")[0]
            else:
                d["UNH_MessageType"] = ""
        if unt:
            d["UNT_SegmentCount"] = unt["elements"][1] if len(unt["elements"]) > 1 else ""
            d["UNT_MessageRef"] = unt["elements"][2] if len(unt["elements"]) > 2 else ""
        if unz:
            d["UNZ_MessageCount"] = unz["elements"][1] if len(unz["elements"]) > 1 else ""
            d["UNZ_ControlReference"] = unz["elements"][2] if len(unz["elements"]) > 2 else ""
    return d


def build_report(pattern_path: str, test_path: str, out_path: str,
                 seg_term: Optional[str] = None, elem_sep: Optional[str] = None,
                 ignore_file: Optional[str] = None, tx: Optional[str] = None):

    ignore_rules = load_ignore_rules(ignore_file)

    pattern_segments, pattern_meta = read_edi_file(pattern_path, seg_term=seg_term, elem_sep_override=elem_sep)
    test_segments, test_meta = read_edi_file(test_path, seg_term=seg_term, elem_sep_override=elem_sep)

    if pattern_meta["format"] != test_meta["format"]:
        raise ValueError("Pattern and test files use different EDI formats (X12 vs EDIFACT).")

    edi_format = pattern_meta["format"]

    pattern_elem_sep = pattern_meta["element_separator"]
    test_elem_sep = test_meta["element_separator"]

    if elem_sep and edi_format == EDI_FORMAT_X12:
        pattern_elem_sep = elem_sep
        test_elem_sep = elem_sep

    p_parsed = parse_segments(pattern_segments, pattern_elem_sep, pattern_meta["release_character"])
    t_parsed = parse_segments(test_segments, test_elem_sep, test_meta["release_character"])

    requested_tx = normalize_tx_type((tx or "").strip(), edi_format) if tx else None

    pattern_detected_raw = detect_tx_type(p_parsed, format_hint=edi_format)
    test_detected_raw = detect_tx_type(t_parsed, format_hint=edi_format)
    pattern_detected = normalize_tx_type(pattern_detected_raw, edi_format) if pattern_detected_raw else None
    test_detected = normalize_tx_type(test_detected_raw, edi_format) if test_detected_raw else None

    if requested_tx:
        tx_type = requested_tx
    else:
        if pattern_detected:
            tx_type = pattern_detected
        elif test_detected:
            tx_type = test_detected
        elif edi_format == EDI_FORMAT_X12:
            tx_type = "856"
        else:
            tx_type = None

    if not tx_type:
        source = "ST01 segments" if edi_format == EDI_FORMAT_X12 else "UNH segments"
        raise ValueError(f"Unable to auto-detect transaction/message type from {source}. Provide --tx to continue.")

    if pattern_detected and pattern_detected != tx_type:
        raise ValueError(f"Transaction type mismatch: pattern file indicates '{pattern_detected_raw}' (normalized to '{pattern_detected}') "
                         f"which differs from expected '{tx_type}'. Use --tx to override if intentional.")
    if test_detected and test_detected != tx_type:
        raise ValueError(f"Transaction type mismatch: test file indicates '{test_detected_raw}' (normalized to '{test_detected}') "
                         f"which differs from expected '{tx_type}'.")

    # Pattern/Test Elements (always emit)
    pattern_elem_values = flatten_elements_for_sheet(p_parsed, "pattern")
    test_elem_values    = flatten_elements_for_sheet(t_parsed, "test")

    seg_diff_rows: List[Dict[str, Any]] = []
    missing_rows: List[Dict[str, Any]] = []
    extra_rows: List[Dict[str, Any]] = []
    elem_diff_rows: List[Dict[str, Any]] = []

    if edi_format == EDI_FORMAT_X12:
        if tx_type == "856":
            compare_856_transactions(p_parsed, t_parsed, ignore_rules,
                                     seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)
        elif tx_type == "861":
            compare_861_transactions(p_parsed, t_parsed, ignore_rules,
                                     seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)
        elif tx_type == "863":
            compare_863_transactions(p_parsed, t_parsed, ignore_rules,
                                     seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)
        else:
            raise ValueError(f"Unsupported X12 transaction type '{tx_type}'. Supported values: 856, 861, 863.")
    elif edi_format == EDI_FORMAT_EDIFACT:
        compare_edifact_transactions(tx_type, p_parsed, t_parsed, ignore_rules,
                                     seg_diff_rows, elem_diff_rows, missing_rows, extra_rows)
    else:
        raise ValueError(f"Unsupported EDI format '{edi_format}'.")

    # Friendly meaning column
    for row in seg_diff_rows:
        if "meaning" not in row:
            row["meaning"] = op_meaning(row["op"])

    # Counts & Summary
    p_counts = Counter([p["tag"] for p in p_parsed])
    t_counts = Counter([t["tag"] for t in t_parsed])
    tag_union = sorted(set(p_counts) | set(t_counts))
    count_rows = [{
        "segment_tag": tag,
        "pattern_count": p_counts.get(tag, 0),
        "test_count": t_counts.get(tag, 0),
        "delta": t_counts.get(tag, 0) - p_counts.get(tag, 0)
    } for tag in tag_union]

    p_env = extract_env(p_parsed, edi_format)
    t_env = extract_env(t_parsed, edi_format)
    summary_rows = [{
        "field": k,
        "pattern": p_env.get(k, ""),
        "test": t_env.get(k, ""),
        "match": "YES" if p_env.get(k, "") == t_env.get(k, "") else "NO"
    } for k in sorted(set(p_env) | set(t_env))]

    # Write Excel
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(count_rows).to_excel(writer, sheet_name="Segment Counts", index=False)

        seg_df = pd.DataFrame(seg_diff_rows)
        if not seg_df.empty:
            seg_cols = ["op", "meaning", "pattern_range", "test_range", "pattern_tag", "test_tag"]
            extra_cols = [c for c in seg_df.columns if c not in seg_cols]
            seg_df[seg_cols + extra_cols].to_excel(writer, sheet_name="Segment Diff", index=False)
        else:
            seg_df.to_excel(writer, sheet_name="Segment Diff", index=False)

        if elem_diff_rows:
            pd.DataFrame(elem_diff_rows)[[
                "line_key","cid_key",
                "pattern_segment_tag","test_segment_tag","element_index_raw",
                "element_position","pattern_value","test_value",
                "status","ignored","status_effective","issue","comments"
            ]].to_excel(writer, sheet_name="Element Diff", index=False)

        if missing_rows:
            pd.DataFrame(missing_rows).to_excel(writer, sheet_name="Missing Segments", index=False)
        if extra_rows:
            pd.DataFrame(extra_rows).to_excel(writer, sheet_name="Extra Segments", index=False)

        pd.DataFrame(pattern_elem_values).to_excel(writer, sheet_name="Pattern Elements", index=False)
        pd.DataFrame(test_elem_values).to_excel(writer, sheet_name="Test Elements", index=False)

    # Styling (autosize + highlighting rules)
    wb = load_workbook(out_path)

    def autosize(name):
        if name in wb.sheetnames:
            ws = wb[name]
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                try:
                    letter = get_column_letter(col[0].column)
                except Exception:
                    continue
                width = max((len(str(c.value)) if c.value else 0) for c in col) + 2
                ws.column_dimensions[letter].width = min(width, 80)

    for sh in ["Summary","Segment Counts","Segment Diff","Element Diff",
               "Missing Segments","Extra Segments","Pattern Elements","Test Elements"]:
        autosize(sh)

    # Element Diff: highlight by status_effective; skip issue/comments; add filter
    if "Element Diff" in wb.sheetnames:
        ws = wb["Element Diff"]
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        header = {cell.value: cell.column for cell in ws[1]}
        status_col = header.get("status_effective") or header.get("status")
        issue_col = header.get("issue")
        comments_col = header.get("comments")
        skip_cols = set(c for c in [issue_col, comments_col] if c)

        if status_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[status_col-1]
                fill = None
                if cell.value in ("MISSING_ELEMENT","EXTRA_ELEMENT"):
                    fill = red
                elif cell.value in ("DIFF","TAG_DIFF"):
                    fill = yellow
                if fill:
                    for col_idx, c in enumerate(row, start=1):
                        if col_idx in skip_cols:
                            continue
                        c.fill = fill

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    # Segment Diff highlighting
    if "Segment Diff" in wb.sheetnames:
        ws = wb["Segment Diff"]
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        header = {cell.value: cell.column for cell in ws[1]}
        op_col = header.get("op")
        if op_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[op_col-1]
                if cell.value in ("INSERT","DELETE","REPLACE_BLOCK"):
                    for c in row:
                        c.fill = red
                elif cell.value == "REPLACE":
                    for c in row:
                        c.fill = yellow

    wb.save(out_path)


# -------------------------------------------------------
#  CLI
# -------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Compare two EDI documents (X12 856/863 or EDIFACT DESADV/QALITY) and export differences to Excel.")
    ap.add_argument("--pattern", required=True, help="Path to the pattern (golden) EDI file")
    ap.add_argument("--test", required=True, help="Path to the test EDI file")
    ap.add_argument("--out", required=True, help="Output Excel file path")
    ap.add_argument("--seg-term", default=None, help="Segment terminator override (optional)")
    ap.add_argument("--elem-sep", default=None, help="Element separator override (X12 only; default auto-detect)")
    ap.add_argument("--ignore-file", default=None, help="CSV file of ignore rules (optional)")
    ap.add_argument("--tx", default=None, help="Transaction/message type (e.g., 856, 863, DESADV, QALITY). Auto-detected when possible.")
    args = ap.parse_args()

    build_report(args.pattern, args.test, args.out,
                 seg_term=args.seg_term, elem_sep=args.elem_sep,
                 ignore_file=args.ignore_file, tx=args.tx)


if __name__ == "__main__":
    main()
